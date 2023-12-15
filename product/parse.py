import psycopg2
from openpyxl import load_workbook
from product.models import Product, Category, Subcategory

# db_params = {
#     'host': 'localhost',
#     'database': 'micasa',
#     'user': 'eren',
#     'password': '1234'
# }

# conn = psycopg2.connect(**db_params)
# cursor = conn.cursor()

def parse_excel():
    excel_file_path = "menu.xlsx"
    sheet_name = "TDSheet"


    wb = load_workbook(excel_file_path)
    sheet = wb[sheet_name]


    data = []

    headers = [cell.value for cell in sheet[2]]


    for row in sheet.iter_rows(min_row=3, min_col=1, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    print(data)
    # try:
    #     for row_data in data:
    #         product_query = "INSERT INTO Product (Nomenclature, Article, Category, Subcategory) " \
    #                         "VALUES (%s, %s, %s, %s);"
    #         cursor.execute(product_query, (row_data['nomenclature'], row_data['article'], 
    #                                     row_data['category'], row_data['subcategory']))
    #         print("Data inserted successfully")
    # except Exception as e:
    #     print(f"Error: {e}")


    for row_data in data:
        category, created = Category.objects.get_or_create(name=row_data['category'])
        sub_category, created = Subcategory.objects.get_or_create(name=row_data['subcategory'], category=category)
        print(category, sub_category, row_data)
        product, created = Product.objects.get_or_create(name=row_data['nomenclature'], article=row_data['article'],
                                                        category=category, subcategory=sub_category)
        

    # conn.commit()
    # cursor.close()
    # conn.close()


